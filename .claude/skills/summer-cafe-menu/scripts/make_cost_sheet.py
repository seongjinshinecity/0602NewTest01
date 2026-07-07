#!/usr/bin/env python3
"""메뉴 데이터 JSON을 받아 원가표 엑셀(cost-sheet.xlsx)을 생성한다.

사용법: python3 make_cost_sheet.py menu-data.json cost-sheet.xlsx

JSON 형식:
{
  "cafe_name": "카페 이름",
  "items": [
    {
      "category": "에이드",
      "name": "블루 하와이 에이드",
      "price": 6500,
      "ingredients": [{"name": "블루큐라소 시럽 30ml", "cost": 450}, ...]
    }
  ]
}
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1B6CA8")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill("solid", fgColor="E8F1F8")
WARN_FILL = PatternFill("solid", fgColor="FDE9E9")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
KRW = '#,##0"원"'
PCT = "0.0%"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    if len(sys.argv) != 3:
        sys.exit("사용법: python3 make_cost_sheet.py <menu-data.json> <output.xlsx>")

    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)

    cafe_name = data.get("cafe_name", "카페")
    items = data["items"]

    wb = Workbook()

    # --- 요약 시트 ---
    ws = wb.active
    ws.title = "원가 요약"
    ws["A1"] = f"{cafe_name} 여름 메뉴 원가표"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["분류", "메뉴명", "판매가", "재료 원가", "원가율", "마진"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    row = 4
    for item in items:
        cost = sum(ing["cost"] for ing in item["ingredients"])
        price = item["price"]
        ratio = cost / price if price else 0
        values = [item["category"], item["name"], price, cost, ratio, price - cost]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = BORDER
        ws.cell(row=row, column=3).number_format = KRW
        ws.cell(row=row, column=4).number_format = KRW
        ws.cell(row=row, column=5).number_format = PCT
        ws.cell(row=row, column=6).number_format = KRW
        if ratio > 0.40:  # 원가율 40% 초과 경고
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # 합계/평균
    n = len(items)
    total_row = row + 1
    ws.cell(row=total_row, column=2, value="평균").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=f"=AVERAGE(C4:C{row-1})").number_format = KRW
    ws.cell(row=total_row, column=4, value=f"=AVERAGE(D4:D{row-1})").number_format = KRW
    ws.cell(row=total_row, column=5, value=f"=AVERAGE(E4:E{row-1})").number_format = PCT
    ws.cell(row=total_row, column=6, value=f"=AVERAGE(F4:F{row-1})").number_format = KRW

    autofit(ws, [10, 24, 12, 12, 10, 12])

    # --- 상세 시트 ---
    ws2 = wb.create_sheet("재료 상세")
    headers2 = ["메뉴명", "재료", "원가"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    r = 2
    for item in items:
        start = r
        for ing in item["ingredients"]:
            ws2.cell(row=r, column=2, value=ing["name"]).border = BORDER
            cell = ws2.cell(row=r, column=3, value=ing["cost"])
            cell.number_format = KRW
            cell.border = BORDER
            r += 1
        name_cell = ws2.cell(row=start, column=1, value=item["name"])
        name_cell.font = Font(bold=True)
        name_cell.fill = CATEGORY_FILL
        if r - 1 > start:
            ws2.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
        for rr in range(start, r):
            ws2.cell(row=rr, column=1).border = BORDER
            ws2.cell(row=rr, column=1).alignment = Alignment(vertical="center")

    autofit(ws2, [24, 28, 12])

    wb.save(sys.argv[2])
    avg_ratio = sum(
        sum(i["cost"] for i in it["ingredients"]) / it["price"] for it in items if it["price"]
    ) / n
    print(f"저장 완료: {sys.argv[2]} (메뉴 {n}개, 평균 원가율 {avg_ratio:.1%})")


if __name__ == "__main__":
    main()
