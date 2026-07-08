# 데일리브루 VoC 분석 엑셀 리포트 생성 — Raw + 분석(수식·차트·피벗·부정Top3)
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

rows = list(csv.reader(open('cafe_reviews.csv', encoding='utf-8')))
header, data = rows[0], rows[1:]
N = len(data)

wb = Workbook()

# ── Sheet 1: Raw 리뷰
ws = wb.active
ws.title = '리뷰원본'
HB = PatternFill('solid', start_color='2F3B52')
HF = Font(bold=True, color='FFFFFF', name='Arial', size=10)
BODY = Font(name='Arial', size=10)
thin = Border(*[Side(style='thin', color='D9D9D9')] * 4)
ws.append(header)
for c in ws[1]:
    c.fill, c.font, c.alignment = HB, HF, Alignment(horizontal='center')
for r in data:
    ws.append([r[0], r[1], int(r[2]), r[3], r[4]])
for row in ws.iter_rows(min_row=2, max_row=N + 1):
    for c in row:
        c.font, c.border = BODY, thin
    row[2].alignment = Alignment(horizontal='center')
# 별점 2 이하 행 강조
NEG = PatternFill('solid', start_color='FDE9E9')
for row in ws.iter_rows(min_row=2, max_row=N + 1):
    if row[2].value <= 2:
        for c in row:
            c.fill = NEG
for col, w in zip('ABCDE', [11, 9, 6, 62, 10]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

# ── Sheet 2: VoC 분석 (전부 수식 기반)
an = wb.create_sheet('VoC분석')
title_font = Font(bold=True, size=13, name='Arial', color='1F2A44')
sec_font = Font(bold=True, size=11, name='Arial', color='2F3B52')
an['A1'] = '데일리브루 손님 리뷰 VoC 분석 (2026-06-22 ~ 07-05, 20건)'
an['A1'].font = title_font

RAW = f"리뷰원본!$C$2:$C${N+1}"   # 별점
THEME = f"리뷰원본!$E$2:$E${N+1}" # 테마

# ① 별점 분포
an['A3'] = '① 별점 분포'
an['A3'].font = sec_font
an['A4'], an['B4'] = '별점', '리뷰 수'
for i, star in enumerate([1, 2, 3, 4, 5]):
    r = 5 + i
    an[f'A{r}'] = star
    an[f'B{r}'] = f'=COUNTIF({RAW},A{r})'
an['A10'] = '합계'
an['B10'] = '=SUM(B5:B9)'
an['A11'] = '평균 별점'
an['B11'] = f'=ROUND(AVERAGE({RAW}),2)'

# ② 테마별 리뷰 수 (피벗)
themes = ['맛', '분위기', '디저트', '대기시간', '가격', '친절', '청결']
an['D3'] = '② 테마별 리뷰 수 (피벗)'
an['D3'].font = sec_font
an['D4'], an['E4'], an['F4'] = '테마', '전체', '부정(≤2점)'
for i, t in enumerate(themes):
    r = 5 + i
    an[f'D{r}'] = t
    an[f'E{r}'] = f'=COUNTIF({THEME},D{r})'
    an[f'F{r}'] = f'=COUNTIFS({THEME},D{r},{RAW},"<=2")'

# ③ 부정 리뷰 Top3 (수식 랭킹)
an['H3'] = '③ 부정 리뷰(별점≤2) 테마 Top3'
an['H3'].font = sec_font
an['H4'], an['I4'], an['J4'] = '순위', '테마', '건수'
for k in range(3):
    r = 5 + k
    an[f'H{r}'] = k + 1
    an[f'J{r}'] = f'=LARGE($F$5:$F$11,{k+1})'
    an[f'I{r}'] = f'=INDEX($D$5:$D$11,MATCH(J{r},$F$5:$F$11,0))'

# ④ 결론 박스
an['A14'] = '④ 핵심 인사이트'
an['A14'].font = sec_font
insights = [
    '긍정 축: 디저트(직접 굽는 치즈케이크·크로플)와 맛 — 5점 리뷰의 중심.',
    '부정 1위: 대기시간 — 부정 리뷰 전부가 주말·점심 웨이팅/품절/회전 문제.',
    '액션: ① 주말 치즈케이크 증량(품절 방지) ② 점심 선포장 테이크아웃 ③ 대기 안내(예상시간) 도입.',
]
for i, s in enumerate(insights):
    an[f'A{15+i}'] = f'• {s}'
    an[f'A{15+i}'].font = BODY

# 헤더 서식 일괄
for cell in ['A4', 'B4', 'D4', 'E4', 'F4', 'H4', 'I4', 'J4']:
    an[cell].fill, an[cell].font = HB, HF
    an[cell].alignment = Alignment(horizontal='center')
for rng in [f'A{r}' for r in range(5, 12)] + [f'B{r}' for r in range(5, 12)]:
    an[rng].border = thin
for r in range(5, 12):
    for col in 'DEF':
        an[f'{col}{r}'].border = thin
for r in range(5, 8):
    for col in 'HIJ':
        an[f'{col}{r}'].border = thin
for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'], [16, 10, 3, 12, 8, 11, 3, 6, 12, 8]):
    an.column_dimensions[col].width = w

# 별점 분포 막대 차트
ch = BarChart()
ch.title = '별점 분포'
ch.y_axis.title = '리뷰 수'
ch.x_axis.title = '별점'
ch.add_data(Reference(an, min_col=2, min_row=4, max_row=9), titles_from_data=True)
ch.set_categories(Reference(an, min_col=1, min_row=5, max_row=9))
ch.height, ch.width = 7, 12
ch.legend = None
an.add_chart(ch, 'A19')

# 테마별 리뷰 수 차트
ch2 = BarChart()
ch2.title = '테마별 리뷰 수 (전체 vs 부정)'
ch2.add_data(Reference(an, min_col=5, min_row=4, max_col=6, max_row=11), titles_from_data=True)
ch2.set_categories(Reference(an, min_col=4, min_row=5, max_row=11))
ch2.height, ch2.width = 7, 14
an.add_chart(ch2, 'H19')

wb.save('voc_report.xlsx')
print('saved voc_report.xlsx')
